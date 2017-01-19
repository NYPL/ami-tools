import argparse
import os
import re
import xlrd
import sys
from openpyxl import load_workbook

excel_profile = [
('reference filename (automatic)', None, None),
('original master', 'bibliographic item', 'title'),
('original master', 'bibliographic item', 'class mark/id'),
('original master', 'bibliographic item', 'division code'),
('original master', 'bibliographic item', 'date'),
('original master', 'object', 'object identifier\n(edit heading to specify type - e.g. barcode)'),
('original master', 'object', 'format'),
('original master', 'object', 'generation'),
('original master', 'object', 'barcode'),
('original master', 'content specifications', 'broadcast\nstandard'),
('original master', 'content specifications', 'color'),
('original master', 'notes', 'condition notes'),
('original master', 'notes', 'content notes'),
('original master', 'notes', 'other notes'),
('file information (automatic)', 'general info', 'filename'),
('file information (automatic)', 'general info', 'extension'),
('file information (automatic)', 'general info', 'file format'),
('file information (automatic)', 'general info', 'duration\n(hh:mm:ss:ff)'),
('file information (automatic)', 'general info', 'date created'),
('file information (automatic)', 'video content', 'video codec name'),
('file information (automatic)', 'audio content', 'audio codec name'),
('operator', 'notes', 'notes')
]


class ExcelValidationError(Exception):
    def __init__(self, value):
        self.value = value
    def __str__(self):
        return repr(self.value)

class AMI_Excel:
    def __init__(self, filename=None):
        self.name = filename
        self.wb = None
        if os.path.exists(filename):
            try:
                self.wb = xlrd.open_workbook(filename)
            except:
                print("not an excel file")
        else:
            print("not a file")

    def validate(self):
        '''
        run each of the validation checks against an AMI Excel Sheet
        '''

        valid = True
        try:
            self.check_sheets()
        except ExcelValidationError as e:
            print("Error in workbook sheets: ", e.value)
            valid = False
        try:
            expected = set([item[0] for item in excel_profile])
            found = set(self.header1)
            self.check_headers_exist(expected, found)
        except ExcelValidationError as e:
            print("Error in first header row: ", e.value)
            valid = False
        try:
            expected = set([item[1] for item in excel_profile if item[1]])
            found = set(self.header2)
            self.check_headers_exist(expected, found)
        except ExcelValidationError as e:
            print("Error in second header row: ", e.value)
            valid = False
        try:
            expected = set([item[2] for item in excel_profile if item[2]])
            found = set(self.header3)
            self.check_headers_exist(expected, found)
        except ExcelValidationError as e:
            print("Error in third header row: ", e.value)
            valid = False
        try:
            self.check_headers_heirarchy(set(excel_profile), set(self.header_heirarchy))
        except ExcelValidationError as e:
            print("Error in header heirarchy: ", e.value)
            valid = False
        try:
            self.check_noequations()
        except ExcelValidationError as e:
            print("Error in cell values: ", e.value)
            valid = False
        return valid

    def check_sheets(self):
        '''
        ensure that Excel contains a sheet named Preservation files
        parse headers if it does
        '''

        for sheet in self.wb.sheet_names():
            if re.match("(preservation|files)", sheet.lower()):
                self.pres_files_sheetname = sheet
                self.pres_files = self.wb.sheet_by_name(sheet)
                self.cols = self.pres_files.ncols
                self.rows = self.pres_files.nrows
                self.parentrow = self.parse_headers()
                break
        else:
            self.raise_excelerror("Required 'Preservation files' sheet is not in workbook.")

        return True

    def parse_headers(self):
        '''
        create lists for each row of headers
        create a list of header heirarchies to parse merged cells
        '''
        self.header1 = self.parse_headerrow(0)
        self.header2 = self.parse_headerrow(1)
        self.header3 = self.parse_headerrow(2)
        self.header_heirarchy = self.parse_headerheirarchy()
        if self.pres_files.cell(0, 0).value.lower() =='reference filename (automatic)':
            self.header_heirarchy.append(('reference filename (automatic)', None, None))


    def parse_headerrow(self, row):
        '''
        create dictionary where keys are header values and values are tuples of column spans
        '''
        headers = []
        for i in range(0, self.cols):
            value = self.pres_files.cell(row, i).value

            if value:
                headers.append(value.lower())

        return headers


    def parse_headerheirarchy(self):
        '''
        create dictionary where keys are header values and values are tuples of column spans
        '''

        header_heirarchy = []

        for i in range(0, self.cols):
            key3 = self.pres_files.cell(2, i).value
            if key3:
                j = i
                key2 = self.pres_files.cell(1, j).value
                while not key2:
                    j -= 1
                    key2 = self.pres_files.cell(1, j).value
                k = j
                key1 = self.pres_files.cell(0, k).value
                while not key1:
                    k-=1
                    key1 = self.pres_files.cell(0, k).value
                header_heirarchy.append((key1.lower(), key2.lower(), key3.lower()))

        return header_heirarchy


    def remove_annoying(self, val1, val2, expected, found):
        '''
        shortcut function to remove items where items are XOR in requirements
        '''

        if ((val1 in expected and val1 in found) and
            (val2 not in found)):
            expected.remove(val2)
        if ((val2 in expected and val2 in found) and
            (val1 not in found)):
            expected.remove(val1)

        return expected


    def check_headers_exist(self, expected_headers, found_headers):
        '''
        validate that a level of headers contains expected values
        '''

        # spreadsheets must have either a barcode field or a object ID field, but both are not required
        expected_headers = self.remove_annoying('barcode',
                                                'object identifier\n(edit heading to specify type - e.g. barcode)',
                                                expected_headers,
                                                found_headers)

        missing_headers = []

        for header in expected_headers:
            if header not in found_headers:
                missing_headers.append(header)

        if missing_headers:
            self.raise_excelerror("Missing required heading - {0}.".format(missing_headers))
        return True


    def check_headers_heirarchy(self, expected, found):
        '''
        validate that each field falls under the merged cells above it
        '''

        # spreadsheets must have either a barcode field or a object ID field, but both are not required
        expected = self.remove_annoying(('original master', 'object', 'barcode'),
                                        ('original master', 'object', 'object identifier\n(edit heading to specify type - e.g. barcode)'),
                                        expected,
                                        found)

        bad_heirarchies = []

        for header in expected:
            if header not in found:
                bad_heirarchies.append(header)

        if bad_heirarchies:
            self.raise_excelerror("Incorrect heirarchy for {0}.".format(bad_heirarchies))
        return True


    def check_noequations(self):
        '''
        validate that all cells are actual values, not formulas
        can't figure out how to do this in xlrd, so let's reload the excel with another library, not best idea
        '''

        self.wb_open = load_workbook(self.name, read_only = True)
        self.pres_files_open = self.wb_open.get_sheet_by_name(self.pres_files_sheetname)

        for i in range(1, self.cols):
            value = self.pres_files_open.cell(row = 4, column = i).value
            # equation check logic, might be better code out there
            if (value and isinstance(value, str) and value[0] == "="):
                self.raise_excelerror("Cell R4C{0} contain equations.".format(i))
        return True


    def raise_excelerror(self, msg):
        '''
        lazy error reporting
        '''

        raise ExcelValidationError(msg)
        logging.error(msg + '\n')
        return False


def _make_parser():
    parser = argparse.ArgumentParser()
    parser.description = "check Excel for validity"
    parser.add_argument("-e", "--excel",
                        help = "path to an AMI Excel file")
    parser.add_argument("-o", "--output",
                        help = "filename to save Excel file if rewritten")
    return parser


def main():
    parser = _make_parser()
    args = parser.parse_args()

    if args.excel:
        print(args.excel)
        excel = AMI_Excel(args.excel)

    if excel and excel.validate():
        print("Validates")
    else:
        print("Does not validate.")
        if args.output:
            wb = load_workbook(args.excel, data_only = True)
            wb.save(args.output)
            new_excel = AMI_Excel(args.output)
            if new_excel.validate():
                print("Validates")
            else:
                print("Does not validate.")



if __name__ == "__main__":
    main()
