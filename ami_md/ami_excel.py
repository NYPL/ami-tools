import os, re, csv, json, datetime, logging

# handling excel
import xlrd
from xlrd import xldate
from openpyxl import load_workbook

# data manipulation and output
import pandas as pd
import numpy as np

# ami modules
import ami_md.ami_md_constants as ami_md_constants
import ami_md.ami_json as ami_json


LOGGER = logging.getLogger(__name__)

class AMIExcelError(Exception):
  def __init__(self, value):
    self.value = value
  def __str__(self):
    return repr(self.value)


class ami_excel:
  def __init__(self, filename):
    """
    Initialize object as excel workbook
    """
    self.path = os.path.abspath(filename)
    self.name = os.path.split(self.path)[1]
    if not os.path.exists(self.path):
      LOGGER.exception("File does not exist at specified path")

    try:
      wb = xlrd.open_workbook(self.path)
    except:
      LOGGER.exception("File is not an excel file")

    self.filename = os.path.splitext(os.path.abspath(filename))[0]
    self.pres_sheet = None
    self.edit_sheet = None
    self.notransfer_sheet = None

    for sheet in wb.sheet_names():
      sheet_lower = sheet.lower()
      #Check if two sheets get identfied by regex below?
      if re.match("(original|preservation|file|full|archive)",
        sheet_lower):
        if not self.pres_sheet:
          self.pres_sheet = ami_pressheet(wb.sheet_by_name(sheet),
            self.name, self.path)
        else:
          raise AMIExcelError("Too many preservation master sheets")
      elif re.match("edit", sheet_lower):
        if not self.edit_sheet:
          self.edit_sheet = ami_editsheet(wb.sheet_by_name(sheet),
            self.name, self.path)
        else:
          raise AMIExcelError("Too many edit master sheets")
      """
      elif re.match("not transferred", sheet_lower):
        self.notransfer_sheet = ami_excelsheet(wb.sheet_by_name(sheet),
          self.path)
      """


  def validate_workbook(self):
    """
    Check the preservation sheet against expectations of Media Ingest.
    """
    valid = True

    LOGGER.info("Checking: {}".format(self.name))
    #Check for a sheet that should have preservation metadata data
    if not self.pres_sheet:
      valid = False
      LOGGER.error("Required sheet for preservation files could not be found in workbook.")

    if not self.pres_sheet.validate_worksheet():
      valid = False
      LOGGER.error("Required sheet for preservation files has errors.")

    return valid


class ami_excelsheet:
  def __init__(self, sheet, wb_name, path):
    """
    Initialize object as excel sheet
    """
    self.path = path
    self.wb = wb_name
    self.name = sheet.name
    self.header_top = self.get_headerRow(sheet, 0)
    self.header_middle = self.get_headerRow(sheet, 1)
    self.header_bottom = self.get_headerRow(sheet, 2)
    self.header_entries = self.get_headerEntries(sheet)
    self.normalized_header_entries = self.get_normalizedHeaderEntries()

    self.sheet_values = pd.read_excel(self.path,
      sheet_name = self.name, skiprows = 2,
      na_values = ami_md_constants.NAS)
    self.sheet_values.columns = self.normalized_header_entries


  def get_headerRow(self, sheet, row):
    """
    Return normalized values from a single row of headers on a
    specified sheet. Newline characters are retained.

    Keyword arguments:
    sheetname -- name of the sheet to extract from (self.pres_sheetname
    or self.edit_sheetname)
    row -- index of the row to extract from (0-2)
    """
    headers = []

    for i in range(0, sheet.ncols):
      value = str(sheet.cell(row, i).value)

      if value:
        headers.append(value)

    return headers


  def get_headerEntries(self, sheet):
    """
    Convenience method to return all header tuples.

    Keyword arguments:
    sheetname -- name of the sheet to extract from (self.pres_sheetname
    or self.edit_sheetname)
    """
    header_entries = []

    for i in range(0, sheet.ncols):
      header_entries.append(self.get_headerEntryAsTuple(sheet, i))

    return header_entries


  def get_headerEntryAsTuple(self, sheet, column):
    """
    Returns tuple of header archiving by stepping backwards from a
    3rd-level header.

    Keyword arguments:
    sheetname -- name of the sheet to extract from (self.pres_sheetname
    or self.edit_sheetname)
    column -- index of the column of the 3rd-level header
    """
    key1, key2, key3 = None, None, None

    key3 = sheet.cell(2, column).value

    j = column
    key2 = sheet.cell(1, j).value
    while not key2:
      j -= 1
      if j == -1:
        key2 = ""
        j = column
        break
      key2 = sheet.cell(1, j).value
    k = column
    key1 = sheet.cell(0, k).value
    while not key1:
      k -= 1
      key1 = sheet.cell(0, k).value

    entry = (str(key1), str(key2), str(key3))

    #Create filename tuple in problematic templates
    if (column == 0 and (not key3 or key1 == '\xa0')):
      entry = ("Reference filename (automatic)", None, None)

    return entry


  def get_normalizedHeaderEntries(self):
    """
    Convert Excel sheet into pandas dataframe
    Each list represents one row of the spreadsheets
    Header values are normalized based on dictionary
    Datetime values are automatically converted to ISO formats
    Row values are normalized based on dictionary

    Keyword arguments:
    sheet -- sheet object from workbook
    """
    normalized_headers = []

    for header_entry in self.header_entries:
      header_entry_list = [entry.lower() if entry else None for entry in header_entry]

      #remove empty tuple values before adding delimiter
      header_string = "|".join(filter(None, header_entry_list))

      #remove newlines since they're inconsistent
      header_string = header_string.replace("\n", " ").strip()
      normalized_header = self.normalize_headerEntry(header_string)
      normalized_headers.append(normalized_header)

    return normalized_headers


  def normalize_headerEntry(self, header_entry):
    """
    Returns a normalized dot-formatted string based on the
    heirarchy of headers from the first three rows of the Excel sheet.

    Keyword arguments:
    header_entry -- string of header values
    """
    if header_entry not in ami_md_constants.HEADER_CONVERSION.keys():
      print(header_entry)

    return ami_md_constants.HEADER_CONVERSION[header_entry]


  def normalize_sheet_values(self):
    """
    Normalize all entries via dictionaries defined in the ami_md_constants
    module. Returns a list of lists
    """
    df = self.sheet_values.dropna(axis = 1, how = "all").astype(object)

    df = df.replace(ami_md_constants.REGEX_REPLACE_DICT, regex=True).astype(object)
    df = df.replace(ami_md_constants.STRING_REPLACE_DICT)

    # add potentially missing, but required information
    if 'source.object.volume' not in df.columns.tolist():
      df['source.object.volumeNumber'] = 1
    df['source.object.volumeNumber'].fillna(1)

    if 'bibliographic.projectCode' not in df.columns.tolist():
      df['bibliographic.projectCode'] = self.wb[0:8]

    if 'asset.fileRole' not in df.columns.tolist():
      df['asset.fileRole'] = df['technical.filename'].str.extract('(em|pm|sc)$', expand=False)

    df['asset.referenceFilename'] = df["technical.filename"] + "." + df["technical.extension"]

    df['source.object.type'] = df['source.object.format'].map(ami_md_constants.FORMAT_TYPE)

    for key in ami_md_constants.MEASURE_UNIT_MAPS.keys():
      value_map = ami_md_constants.MEASURE_UNIT_MAPS[key]
      df = self.map_formatvalue(df,
        value_map['from_column'],
        value_map['to_column'],
        value_map['constant_value'],
        value_map['values_map_column'],
        value_map['values_map'])

    if 'bibliographic.primaryID' not in df.columns.tolist():
      df = self.map_primaryid(df)

    #force all the numerics back to numeric, and drop all empty columns
    df.sort_index(axis=1, inplace=True)

    self.sheet_values = df

    return


  def map_formatvalue(self, df, from_column, to_column, value = None,
    values_map_column = None, values_map = None):
    """
    Conditionally map units for columns with measures

    Keyword arguments:
    df -- pandas dataframe to operate on
    from_column -- name of column to map from
    to_column -- name of column to map new values to
    value -- constant value to map into to_column
    values_map_column -- name of column to determin variable mapping
    values_map -- dictionary of mapping values
    """
    if from_column not in df.columns:
      return df

    df[from_column] = pd.to_numeric(df[from_column], errors='coerce')

    if value:
      df[to_column] = np.where(df[from_column].notnull(), value, None)
    elif values_map_column and values_map:
      #add unit value regardless if there is a measure value
      df[to_column] = df[values_map_column].map(values_map)
      #reset all unit values where there's no corresponding measure
      df[to_column] = df[to_column].mask(df[from_column].isnull(), None)

    return df


  def map_primaryid(self, df):
    primary_id_columns = ["bibliographic.cmsItemID",
                          "bibliographic.classmark",
                          "bibliographic.barcode"]
    for column in primary_id_columns:
      if column in df.columns.tolist():
        has_id = df[column].notnull()
        df.loc[has_id, 'bibliographic.primaryID'] = df[column]
        break

    return df


  def convert_amiExcelToCSV(self, csv_path, normalize = True):
    """
    Convert a single Excel sheet into a CSV with normalized contents.

    Keyword arguments:
    sheetname -- name of the sheet to convert (self.pres_sheetname or
    self.edit_sheetname)
    csv_filename -- path of the output file
    """
    if normalize:
      self.normalize_sheet_values()

    LOGGER.info("Writing {}".format(csv_path))
    self.sheet_values.to_csv(csv_path, index = False)


  def convert_amiExcelToJSON(self, json_directory,
    schema_version = "x.0.0", filepaths = None):
    """
    Convert all rows in an Excel sheet into JSON files with
    normalized data. Filename is based on described file's name.

    Keyword arguments:
    sheetname -- name of the sheet to convert (self.pres_sheetname or
    self.edit_sheetname)
    json_directory -- path to output directory for json files
    """
    self.normalize_sheet_values()

    json_directory = os.path.abspath(json_directory)
    df = self.sheet_values

    if filepaths:
      for filepath in filepaths:
        media_filename =  os.path.splitext(os.path.basename(filepath))[0]

        try:
          row = df[df["technical.filename"] == os.path.splitext(media_filename)[0]]
        except:
          raise_excelerror("Excel sheet does not have a record for {}".format(media_filename))

        row_dict = row.squeeze().to_dict()
        row_dict["asset.referenceFilename"] = media_filename

        json_tree = ami_json.ami_json(flat_dict = row_dict,
          filepath = filepath, load = False,
          schema_version = schema_version, media_filepath = os.path.splitext(filepath)[0])
        json_tree.repair_techmd()
        json_tree.write_json(json_directory)

    else:
      for (index, row) in self.sheet_values.iterrows():
        json_tree = ami_json.ami_json(flat_dict = row.to_dict(),
          schema_version = schema_version)
        json_tree.write_json(json_directory)


  def raise_excelerror(self, msg):
    """
    lazy error reporting
    """
    LOGGER.error(msg)
    return False
    raise AMIExcelError(msg)



class ami_pressheet(ami_excelsheet):

  def __init__(self, *args, **kwargs):
    super(ami_pressheet, self).__init__(*args, **kwargs)

  def validate_worksheet(self):
    """
    Check the preservation sheet against expectations of Media Ingest.
    """
    valid = True

    #Check that sheet contains required headers
    for i in range(0, 3):
      try:
        expected = set([item[i] for item in ami_md_constants.MEDIAINGEST_EXPECTED_HEADERS if item[i]])
        found = set([item[i] for item in self.header_entries if item[i]])
        self.check_headerRow(expected, found)
      except AMIExcelError as e:
        print("Error in header row of sheet {}: {}"
          .format(self.name, e.value))
        valid = False

    #Check that sheet headers have the correct heirarchy
    try:
      header_entries = set(self.header_entries)
      self.check_headerEntries(
        set(ami_md_constants.MEDIAINGEST_EXPECTED_HEADERS),
        header_entries)
    except AMIExcelError as e:
      print("Error in header entries: ", e.value)
      valid = False

    #Check that Reference Filename field actually exists
    try:
      self.check_reffilenameheader()
    except AMIExcelError as e:
      print("Missing R1C1 header: Reference filename (automatic)")
      valid = False

    #Check that the preservation sheet does not contain equations
    try:
      self.check_noequations()
    except AMIExcelError as e:
      print("Error in cell values: ", e.value)
      valid = False

    return valid


  def remove_annoying(self, val1, val2, expected, found):
    """
    Convenience function to remove items with XOR requirements

    Keyword arguments:
    val1 -- first XOR item
    val2 -- seocnd XOR item
    expected -- set of all required items
    found -- set of items that were extracted from sheet
    """
    if ((val1 in expected and val1 in found) and
      (val2 not in found)):
      expected.remove(val2)

    if ((val2 in expected and val2 in found) and
      (val1 not in found)):
      expected.remove(val1)

    return expected


  def check_headerRow(self, expected, found):
    """
    Check that a row of headers contains all the required values

    Keyword arguments:
    expected -- set of all required header values
    found -- set of items that were extracted from sheet
    """
    # spreadsheets must have either a barcode field or a object ID field
    # but both are not required
    header1 = 'Barcode'
    header2 = ('Object identifier\n(edit heading to specify type' +
      ' - e.g. barcode)')
    expected = self.remove_annoying(header1, header2, expected, found)

    missing = []

    for header in expected:
      if header not in found:
        missing.append(header)

    if missing:
      raise AMIExcelError("Missing required value- {0}."
        .format(missing))

    return True


  def check_headerEntries(self, expected, found):
    """
    Check that a sheet contains the required heirarchy of headers

    Keyword arguments:
    expected -- set of all required header heirarchies values
    found -- set of header heirarchies that were extracted from sheet
    """
    # spreadsheets must have either a barcode field or a object ID field
    # but both are not required
    header1 = ('Original master', 'Object', 'Barcode')
    header2 = ('Original master', 'Object',
      'Object identifier\n(edit heading to specify type ' +
      '- e.g. barcode)')
    expected = self.remove_annoying(header1, header2, expected, found)

    bad_entries = []

    for header in expected:
      if header not in found:
        bad_entries.append(header)

    if bad_entries:
      raise AMIExcelError("Incorrect header entry for {0}."
        .format(bad_entries))

    return True

  def check_reffilenameheader(self):
    if 'Reference filename (automatic)' not in self.header_top:
      raise AMIExcelError("Missing correct filename header.")

    return True


  def check_noequations(self):
    """
    Verify that no column in a sheet contains an equation
    Based on checking every cell in the 4th row

    Keyword arguments:
    sheet -- sheet object from workbook
    """
    wb_open = load_workbook(self.path, read_only = True)
    sheet = wb_open[self.name]

    for i in range(1, len(self.header_entries)):
      for j in range(1,5):
        value = sheet.cell(row = j, column = i).value
        # equation check logic, TODO might be better code out there
        if (value and isinstance(value, str) and value[0] == "="):
          raise AMIExcelError("Cell R4C{0} contain equations."
            .format(i))

    return True


class ami_editsheet(ami_excelsheet):

  def __init__(self, *args, **kwargs):
    super(ami_editsheet, self).__init__(*args, **kwargs)

  def add_PMDataToEM(self, pm_data):
    """
    """
    em_df = self.sheet_values
    em_df["join_idx"] = em_df["technical.filename"].str.slice(0, -3)

    pm_df = pm_data.copy()
    pm_drop_cols = set(pm_df.columns.tolist()).intersection(set(em_df.columns.tolist()))
    pm_df["join_idx"] = pm_df["technical.filename"].str.slice(0, -3)
    #pm_drop_cols = [col for col in pm_df.columns.tolist() if (col.startswith('technical') or col.startswith('digitizer'))]
    pm_df = pm_df.drop(pm_drop_cols, axis = 1)

    em_df = em_df.join(pm_df.set_index("join_idx"), on = "join_idx")
    em_df = em_df.drop("join_idx", axis = 1)
    em_df["asset.referenceFilename"] = em_df["technical.filename"] + "." + em_df["technical.extension"]

    self.sheet_values = em_df
